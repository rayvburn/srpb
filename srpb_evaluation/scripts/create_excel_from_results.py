'''
Creates an Excel file with results (raw and filtered) obtained from SRPB logs.
Note that CSV with results per each run must be previously generated. Otherwise the script will notify that
results file is missing.

Dependency:
sudo apt install python3-openpyxl
'''

import csv
import excel_sheet_defines
import glob
import re
import sys

from string import ascii_uppercase
from openpyxl import Workbook
from pathlib import Path
from typing import List
from typing import Dict


def get_log_dirs_planner(dir_path: str, planner_name: str, min_logs=3):
    dirnames_underscore = glob.glob(dir_path + '/' + '*_' + planner_name + '*')
    dirnames_dash = glob.glob(dir_path + '/' + '*-' + planner_name + '*')
    dirnames = dirnames_underscore
    dirnames.extend(dirnames_dash)
    if len(dirnames) < min_logs:
        exit(f'Too few data entries for {planner_name} planner. Got {len(dirnames)}/{min_logs}: \r\n{dirnames}')

    # got rid of the dirs that are marked to be ignored
    dirnames_valid = []
    for dirname in dirnames:
        if re.search('ignore', dirname, re.IGNORECASE):
            print(f'Ignoring a directory `{dirname}` from including to the overall `{planner_name}` planner results')
            continue
        dirnames_valid.append(dirname)

    return dirnames_valid


def get_log_result_files(log_dirs: str):
    result_files = []
    for log_dir in log_dirs:
        results_file = glob.glob(log_dir + '/' + '*results.txt*')
        if len(results_file) == 0:
            print(f'Lack of results file at `{log_dir}` dir, skipping')
            continue
        elif len(results_file) > 1:
            print(f'Multiple results files at `{log_dir}` dir, skipping')
            continue
        result_files.append(results_file[0])
    return result_files


def read_results_file(results_file: str):
    with open(results_file, newline='') as csvfile:
        csv_result = csv.reader(csvfile, delimiter=',', quotechar='|')
        # file path only last 2 levels
        file = Path(results_file).parts[-2] + '/' + Path(results_file).parts[-1]
        # delete part
        file = str(file).rstrip('_results.txt')
        dict = {}
        dict['file'] = file
        for row in csv_result:
            key = str(row[0]).lstrip().rstrip()
            value_str = str(row[1])
            try:
                if 'nan' in value_str:
                    value = None
                    print(f"NaN value detected for the key `{key}` in the results file `{results_file}`")
                elif value_str.find('.') != -1:
                    value = float(value_str)
                else:
                    value = int(value_str)
            except ValueError:
                print(f"Cannot convert a value `{value_str}` of a key `{key}` to a numeric, skipping `{results_file}`")
                return {}
            dict[key] = value
        return dict


def collect_results(dir_path: str, planner_name: str):
    log_dirnames = get_log_dirs_planner(dir_path, planner_name)
    results_files = get_log_result_files(log_dirnames)
    # list of dicts with results
    results_total = []
    for result_file in results_files:
        results = read_results_file(result_file)
        results_total.append(results)
    return results_total


def collect_results_planners(dir_path: str, planner_name: List[str]) -> Dict:
    results_planners = {}
    for name in planner_name:
        results_planners[name] = collect_results(dir_path, name)
    return results_planners


def prepare_sheet_rows(results_total: Dict) -> List[List]:
    planners_row = ['Planner']
    trials_row = ['Trial']
    file_row = ['File']
    spls_robot_row = ['Samples robot']
    spls_ppl_row = ['Samples people']
    spls_grp_row = ['Samples groups']
    m_goal_row = ['m_goal']
    m_obs_row = ['m_obs']
    m_eff_row = ['m_eff']
    m_cef_row = ['m_cef']
    m_cre_row = ['m_cre']
    m_vsm_row = ['m_vsm']
    m_hsm_row = ['m_hsm']
    m_path_row = ['m_path']
    m_chc_row = ['m_chc']
    m_osc_row = ['m_osc']
    m_bwd_row = ['m_bwd']
    m_irot_row = ['m_irot']
    m_psi_row = ['m_psi']
    m_fsi_row = ['m_fsi']
    m_dir_row = ['m_dir']
    m_psd_row = ['m_psd']

    # iterate over keys
    for planner_key in results_total:
        for i, result in enumerate(results_total[planner_key]):
            planners_row.append(planner_key)
            trials_row.append(i + 1)
            file_row.append(result['file'])
            spls_robot_row.append(result['s_rbt'])
            spls_ppl_row.append(result['s_ppl'])
            spls_grp_row.append(result['s_grp'])
            m_goal_row.append(result['m_goal'])
            m_obs_row.append(result['m_obs'])
            m_eff_row.append(result['m_mef'])
            m_cef_row.append(result['m_cef'])
            m_cre_row.append(result['m_cre'])
            m_vsm_row.append(result['m_vsm'])
            m_hsm_row.append(result['m_hsm'])
            m_path_row.append(result['m_path'])
            m_chc_row.append(result['m_chc'])
            m_osc_row.append(result['m_osc'])
            m_bwd_row.append(result['m_bwd'])
            m_irot_row.append(result['m_inp'])
            m_psi_row.append(result['m_psi'])
            m_fsi_row.append(result['m_fsi'])
            m_dir_row.append(result['m_dir'])
            m_psd_row.append(result['m_psd'])

    return [
        planners_row,
        trials_row,
        file_row,
        spls_robot_row,
        spls_ppl_row,
        spls_grp_row,
        m_goal_row,
        m_obs_row,
        m_eff_row,
        m_cef_row,
        m_cre_row,
        m_vsm_row,
        m_hsm_row,
        m_path_row,
        m_chc_row,
        m_osc_row,
        m_bwd_row,
        m_irot_row,
        m_psi_row,
        m_fsi_row,
        m_dir_row,
        m_psd_row
    ]

def cell_coords_to_sheet_cell_id(row: int, col: int):
    row_w_offset = row + 1 + 2
    col_w_offset = col + 1
    cell_row = str(row_w_offset)

    found_col = False
    it_col = 0
    cell_col = ''
    for c in ascii_uppercase:
        if it_col == col_w_offset:
            cell_col = str(c)
            found_col = True
            break
        it_col = it_col + 1
    # repeat once again if required (doubling letters) - columns A - ZZ cover a sufficient number of cases
    if not found_col:
        # first letter of the cell address
        for first in ascii_uppercase:
            # second letter of the cell address
            for second in ascii_uppercase:
                # print(f'[cell_coords_to_sheet_cell_id] row {row}, col {col}, cell row {cell_row} | it col {it_col}, ADDR `{first}{second}`')
                if it_col == col_w_offset:
                    cell_col = str(first) + str(second)
                    found_col = True
                    break
                it_col = it_col + 1
            # break the outer loop if possible
            if found_col:
                break
    if not found_col:
        exit(f"Could not find a valid column ID for ({row}, {col}) configuration")

    # append row
    cell_id = str(cell_col) + str(cell_row)
    return cell_id


def get_sheet_datacell(planner: str, trial: int, result_key: str, results_total: Dict):
    # we must have some data
    if len(results_total) == 0:
        exit('Cannot proceed as results are empty')
    if not planner in results_total.keys():
        exit(f'Cannot proceed as results do not contain that planner: {planner}. Available keys: {results_total.keys()}')

    # counting overall number of trials (to determine column)
    trial_counter = 0
    found = False
    for planner_key in results_total:
        for i, result in enumerate(results_total[planner_key]):
            if i == trial and planner == planner_key:
                found = True
                break
            trial_counter = trial_counter + 1
        if found:
            break

    # counting rows from the start - exemplary key to iterate through metrics
    metric_counter = 0
    try:
        results_example = results_total[planner][0]
    except IndexError:
        print(
            f'The planner `{planner}` does not seem to contain any keys with metric names. '
            f'Got `{results_total[planner]}` while searching for the metric `{result_key}` for the trial `{trial}`'
        )
        return None

    if not result_key in results_example.keys():
        exit(f'Cannot proceed as results do not contain that key: `{result_key}`. Available keys: `{results_example.keys()}`')

    found = False
    for key in results_example:
        if result_key == key:
            found = True
            break
        metric_counter = metric_counter + 1

    return cell_coords_to_sheet_cell_id(metric_counter, trial_counter)


# Works for A-Z range
def increment_char(c: chr):
    return chr(ord(c) + 1)


def calculate_sheet(wb: Workbook, planner_names: List[str], results_total: Dict, calc_fun='MEDIAN'):
    row_start = int(excel_sheet_defines.RESULT_INIT_ROW)
    col_header = str(excel_sheet_defines.RESULT_INIT_COL)
    ws[col_header + str(row_start)] = 'Planner'
    ws[col_header + str(row_start + 1)] = 'Trials'
    row_metric_start = row_start + 1
    col_planner_start = str(increment_char(col_header))
    col_planner = col_planner_start

    # retrieve metric keys from an arbitrary planner
    metric_keys = []
    results_example = results_total[planner_names[0]][0]
    for key in results_example.keys():
        metric_keys.append(str(key))

    # create compound representation (filtered) of raw data
    for planner in planner_names:
        # investigate each metric
        for m_index, metric in enumerate(metric_keys):
            # don't need that
            if metric == 'file':
                continue
            # each trial of a specific planner: find range of cells
            planner_trials = len(results_total[planner])
            planner_trial_last_id = planner_trials - 1
            cell_begin = get_sheet_datacell(planner, 0, metric, results)
            cell_end = get_sheet_datacell(planner, planner_trial_last_id, metric, results)
            if cell_begin == None or cell_end == None:
                continue
            # fill spreadsheet
            ws[col_header  + str(row_metric_start + m_index)] = metric
            ws[col_planner + str(row_start)] = planner
            ws[col_planner + str(row_start + 1)] = planner_trials
            # filtering: put excel function here
            ws[col_planner + str(row_metric_start + m_index)] = '=' + str(calc_fun) + '(' + str(cell_begin) + ':' + str(cell_end) + ')'
        # next planner
        col_planner = increment_char(col_planner)


########################################################
# command line arguments
if len(sys.argv) == 1:
    print(f'Usage: ')
    print(f'')
    print(f'  python3 {sys.argv[0]}  <path to the main directory with SRPB logs>  <(OPTIONAL) space-separated planner names>')
    print(f'')
    print(f'  1. The main directory with logs should store the log files grouped into separate directories, ')
    print(f'     each containing a robot, people and group data.')
    print(f'  2. In order to detect which planners to include into results sheet, each name of a planner ')
    print(f'     must be a part of log filenames.')
    print(f'')
    exit()

logs_dir = str(sys.argv[1])
print(f'Name of directory with logs: {logs_dir}')

planners = []
if not len(sys.argv) == 2:
    planners = sys.argv[2:]
else:
    planners = ['teb', 'dwa', 'trajectory', 'eband', 'hateb', 'cohan']

print(f'Investigated planners: {planners}')
results = collect_results_planners(logs_dir, planners)

###############################################
# documentation at https://openpyxl.readthedocs.io/en/stable/index.html
wb = Workbook()

# grab the active worksheet (workbook always creates an arbitrary sheet)
ws_init = wb.active
# create a sheet with a custom name
ws = wb.create_sheet(title=excel_sheet_defines.SHEET_NAME)
# delete the sheet initially created
wb.remove_sheet(worksheet=ws_init)

# Prepare rows of data and append them to the sheet
rows = prepare_sheet_rows(results)
for row in rows:
    ws.append(row)

# make some calculations
calculate_sheet(ws, planners, results, 'MEDIAN')

# Prepare name of the output file
output_filename = 'results' + '_' + Path(logs_dir).name
for planner in planners:
    output_filename = output_filename + '_' + planner
output_filename = output_filename.rstrip('_') + '.xlsx'
output_path = Path(logs_dir).parent.absolute() / output_filename

# Save the file
wb.save(str(output_path))

print(f'Results saved in: {output_path}')
print("")
# When the sheet with the results is not opened and saved by the Excel or LibreOffice Calc, then reading a non-empty
# cell will probably return None
# Ref1: https://itecnote.com/tecnote/python-openpyxl-data_onlytrue-returning-none/
# Ref2: https://groups.google.com/g/openpyxl-users/c/GbBOnOa8g7Y
print(f'Consider opening the results file and saving it with Excel/LibreOffice Calc (without any modifications).')
print(f'It will produce cached values based on formulas written (`openpyxl` library is not able to do so).')
print(f'This is a necessary step when one wants to use the script that creates a LaTeX table from a spreadsheet')
